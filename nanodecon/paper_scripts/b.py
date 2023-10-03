import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

# Create a new Excel workbook
wb = openpyxl.Workbook()

# Select the default worksheet
ws = wb.active

# Set the title and subtitle
ws['A1'] = 'Melbourne 3 months'
ws['A2'] = 'My Name'

# Create the expense table
expenses = {
    'Flight': 10000,
    'Housing': 30000,
    'Transport': 2000,
    'Kost': 7500,
    'Insurance': 2000,
}

# Populate the expense table
ws['A4'] = 'Expense'
ws['B4'] = 'Amount'

for i, (expense, amount) in enumerate(expenses.items(), start=5):
    ws[f'A{i}'] = expense
    ws[f'B{i}'] = amount

# Create the income table
ws['D4'] = 'Income'
ws['E4'] = 'Amount'
ws['D5'] = 'Salary'
ws['E5'] = 58200

# Calculate and display the total sum
total_expenses_formula = f'=SUM(B5:B{4 + len(expenses)})'
total_income_formula = f'E5'
total_sum_formula = f'=E7-B7'

ws['A7'] = 'Total'
ws['B7'] = total_expenses_formula
ws['D7'] = 'Total'
ws['E7'] = total_income_formula
ws['E8'] = total_sum_formula

# Apply bold font to the total row
for cell in ws['A7:E7']:
    for row in cell:
        row.font = Font(bold=True)

# Save the Excel file
wb.save('budget.xlsx')

